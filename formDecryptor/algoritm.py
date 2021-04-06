from io import BytesIO

import cv2
import numpy as np
from PIL import Image
from django.core.files.base import ContentFile
from django.utils.translation import gettext_lazy

from formDecryptor.models import Form
from mainApp.models import CSVFile

from openpyxl import Workbook
#from openpyxl.styles import Color, Fill, Font, PatternFill, Border, Side
#from openpyxl.cell import Cell

class Algoritm:
    def __init__(self, bukvi, otveti, radW, radH, stolb, file_count, file_list, user):
        self.bukvi = sorted((str.upper(bukvi)).split(), key=str.upper, reverse=True)
        self.otveti = otveti.split(" ")
        for i in range(len(self.otveti)):
            self.otveti[i] = str.upper(self.otveti[i])

        self.radW = radW
        self.radH = radH
        self.stolb = stolb
        self.klass = 11

        self.otvet = []
        self.otvetb = []
        self.balli = []
        self.oshibki = []

        self.user = user
        self.file_list = file_list
        self.count = file_count

    def __index__(self):
        return self.create_excel()

    def create_numbers(self, *args, **kwargs):
        v = 1
        numbers = []
        numbers.append("Имя файла")
        numbers.append("")
        numbers.append('Класс')
        numbers.append('')
        while v < kwargs['count'] + 1:
            numbers.append(v)
            v += 1
        numbers.append('Итого')
        return numbers

    def get_main_color(self, image):
        colors = image.getcolors(256)
        max_occurence, most_present = 0, 0
        try:
            for c in colors:
                if c[0] > max_occurence:
                    (max_occurence, most_present) = c
            return most_present
        except TypeError:
            raise Exception("Too many colors in the image")

    def main_work(self, W, H, stolb, nachalo, image, sto, mini, maxi):
        vhod = W * stolb
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        gray = cv2.bitwise_not(gray)

        thresh = cv2.threshold(gray, 30, 255, cv2.THRESH_BINARY)[1]

        contours = cv2.findContours(thresh, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        contours = contours[0] if len(contours) == 2 else contours[1]

        result = image.copy()

        X = []
        Y = []

        min_area = mini
        max_area = maxi

        for c in contours:

            area = cv2.contourArea(c)
            if area > min_area and area < max_area:
                rect = cv2.minAreaRect(c)
                box = cv2.boxPoints(rect)
                box = np.int0(box)
                cv2.drawContours(result, [box], -1, (255, 0, 0), 1)

        s = 0
        for g in box:
            X.append(tuple(box[s])[0])
            Y.append(tuple(box[s])[1])
            s = s + 1

        y = min(Y)
        x = min(X)
        w = max(X) - min(X)
        h = max(Y) - min(Y)

        crop_image = image[y:y + h, x:x + w]
        graYP = cv2.cvtColor(crop_image, cv2.COLOR_BGR2GRAY)
        graYP = cv2.bitwise_not(graYP)
        thresh2 = cv2.threshold(graYP, 30, 255, cv2.THRESH_BINARY)[1]

        contours = cv2.findContours(thresh2, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        contours = contours[0] if len(contours) == 2 else contours[1]

        min_area1 = 80
        max_area1 = 140

        n = 0
        s = 0
        first = 0
        old = 0
        oldY = 0

        XP = []
        YP = []

        for c in range(nachalo, W):
            XP.append([])
        for c in range(H):
            YP.append([])

        for c in contours:
            area = cv2.contourArea(c)
            if area > min_area1 and area < max_area1:
                rect = cv2.minAreaRect(c)
                box = cv2.boxPoints(rect)
                box = np.int0(box)
                X1 = []
                Y1 = []
                for l in range(4):
                    X1.append(list(box[l])[0])
                    Y1.append(list(box[l])[1])

                XO = max(X1) - min(X1)
                YO = max(Y1) - min(Y1)
                if (old - 5 > (box[0])[0] or old + 5 < (box[0])[0]) or (
                        oldY - 5 > (box[0])[1] or oldY + 5 < (box[0])[1]):
                    if ((XO - 5 < YO and XO + 5 > YO) or (XO > YO - 5 and XO < YO + 5)):
                        if s == 0 and sto == 0:
                            first = round((max(Y1) + min(Y1)) / 2)
                        if s == W:
                            s = 0
                            n += 1
                        for l in range(round(((max(Y1) + min(Y1)) / 2)) - 3, round(((max(Y1) + min(Y1)) / 2)) + 3):
                            if l == first or sto == 1:
                                if s > W - nachalo - 1 and sto == 0:
                                    XP[W - nachalo - 1].append(round((max(X1) + min(X1)) / 2))
                                else:
                                    XP[s].append(round((max(X1) + min(X1)) / 2))
                                YP[n * sto].append(round((max(Y1) + min(Y1)) / 2))
                                break

                        s += 1

                old = (box[0])[0]
                oldY = (box[0])[1]
        XP = sorted(XP, key=None, reverse=True)

        X2 = []
        Y2 = []

        for c in range(W - nachalo):
            XP[c] = sorted(XP[c])
        for c in range(H):
            YP[c] = sorted(YP[c])

        n = 0
        bold = 0
        for c in XP:
            kolv = 0
            bold = 0
            vse = {}

            for b in c:
                if bold != b:
                    kolv = 0
                kolv += 1
                vse[b] = kolv
                bold = b

            bold = 0
            n = max(vse.values())
            for b in c:
                if bold != b:
                    if vse[b] == n:
                        if b > bold + 5:
                            X2.append(b)

                if vse[b] == n:
                    bold = b

        for c in YP:
            kolv = 0
            bold = 0
            vse = {}

            for b in c:
                if bold != b:
                    kolv = 0
                kolv += 1
                vse[b] = kolv
                bold = b

            bold = 0
            n = max(vse.values())
            for b in c:
                if bold != b:
                    if vse[b] == n:
                        if b > bold + 5:
                            Y2.append(b)

                if vse[b] == n:
                    bold = b

        X2 = sorted(set(X2), key=None, reverse=True)
        Y2 = sorted(set(Y2), key=None, reverse=True)
        Aold = 0
        if sto == 0:
            for A in X2:
                if Aold < A + 5:
                    X2.remove(A)
                Aold = A

        n = 0
        s = 1
        kolv = 0
        old = 0
        oldY = 0

        for c in contours:

            area = cv2.contourArea(c)
            if area > min_area1 and area < max_area1:
                rect = cv2.minAreaRect(c)
                box = cv2.boxPoints(rect)
                box = np.int0(box)
                cv2.drawContours(crop_image, [box], -1, (255, 0, 0), 1)
                X1 = []
                Y1 = []
                for l in range(4):
                    X1.append(tuple(box[l])[0])
                    Y1.append(tuple(box[l])[1])

                XO = max(X1) - min(X1)
                YO = max(Y1) - min(Y1)
                if (old - 5 > (box[0])[0] or old + 5 < (box[0])[0]) or (
                        oldY - 5 > (box[0])[1] or oldY + 5 < (box[0])[1]):
                    if ((XO - 6 < YO and XO + 6 > YO) or (XO > YO - 6 and XO < YO + 6)):
                        cv2.drawContours(crop_image, [box], -1, (255, 0, 0), 1)

                        y = min(Y1) + 1
                        x = min(X1) + 1
                        w = max(X1) - min(X1) - 2
                        h = max(Y1) - min(Y1) - 2

                        cube = thresh2[y:y + h, x:x + w]
                        V = self.get_main_color(Image.fromarray(np.uint8(cube)))

                        kolv += 1
                        n += 1
                        if n > W * (H / 2):
                            s += 1
                            n = 0

                        if V != 0:
                            rub = 0
                            for g in X2:
                                if ((max(X1) + min(X1)) / 2) + 10 > g:
                                    break
                                rub += 1

                            rubs = 0
                            if sto == 1:
                                for g in Y2:
                                    if ((max(Y1) + min(Y1)) / 2) + 10 > g:
                                        break

                                    rubs = rubs + 1
                                    if rubs >= (H / stolb):
                                        rubs = 0

                            if sto == 1:
                                self.otvet.append((round(((vhod) / s) - (rub))))
                                self.otvetb.append(self.bukvi[rubs])

                            if sto == 0:
                                for l in range(round(((max(Y1) + min(Y1)) / 2)) - 4,
                                               round(((max(Y1) + min(Y1)) / 2)) + 4):
                                    if l == Y2[0]:
                                        self.otvet.append(round((vhod - rub - 1)))
                                        self.otvetb.append("klass")
                                        break

                            cv2.circle(crop_image, (((c[0])[0])[0], ((c[0])[0])[1]), 5, (0, 0, 255), -1)
                old = (box[0])[0]
                oldY = (box[0])[1]

        if kolv != (W * H) - nachalo:
            raise Exception

    def create_excel(self, *args, **kwargs):
        a12 = 0
        bytes = BytesIO()

        wb = Workbook()
        wr = wb.active

        s = 1

        # while s <= radW*stolb + 6:
        #    wr.cell(column=1, row=s).font = Font(bold=True, size=10)
        #    wr.cell(column=1, row=s).fill = PatternFill(bgColor='A9A9A9')
        #    s += 1

        wr.append(self.create_numbers(count=self.radW * self.stolb))
        self.oshibki.clear()

        while int(a12) < int(self.count) - 1:
            try:
                self.otvet.clear()
                self.otvetb.clear()
                self.balli.clear()
                pk = int(self.file_list[str(a12)])
                print(pk)
                form = Form.objects.get(pk=pk)
                img = cv2.imread(form.get_full_url())

                self.main_work(self.radW, self.radH, self.stolb, 0, img, 1, 75000, 100000)
                self.main_work(self.klass, 1, 1, 1, img, 0, 50000, 80000)

                klass1 = self.otvet[-1]
                del self.otvet[-1]
                del self.otvetb[-1]

                for i in range(1, self.radW * self.stolb + 1):
                    if i not in self.otvet:
                        self.otvet.append(i)
                        self.otvetb.append(" ")

                konez = np.c_[self.otvet, self.otvetb]
                konez = konez[konez[:, 0].astype(int).argsort()]

                b = 0

                bold = 0

                for i in range(len(konez)):
                    a = str((konez[i])[0]) + (konez[i])[1]
                    b = (konez[i])[0]
                    if b != bold:
                        if a in self.otveti:
                            self.balli.append(1)
                        else:
                            self.balli.append(0)
                    else:
                        self.balli[i - 1] = 0
                        self.balli.append(0)
                    bold = b

                konez = np.c_[konez, self.balli]
                vsegoballov = sum(self.balli)

                for q in self.check_array(form.get_file_name(), konez, klass1, vsegoballov):
                    wr.append(q)
                wr.append([])

                a12 += 1
            except Exception as e:
                pk = int(self.file_list[str(a12)])
                form = Form.objects.get(pk=pk)
                print(pk)
                self.oshibki.append("Диффектный документ: " + str(form.get_file_name()))
                a12 += 1

        wb.save(bytes)
        file = ContentFile(bytes.getvalue())

        final_excel = CSVFile(user=self.user)
        try:
            final_excel.save()
            final_excel.file.save('Data_of_olympiads.xlsx', file)
        except Exception as e:
            return [None, [gettext_lazy(e.__str__())]]

        return [final_excel, self.oshibki]

    def check_array(self, file_name, array, klass, itog):
        l = 1
        result = [[], []]

        result[0].append(file_name)
        result[0].append('')

        result[0].append(klass)
        result[0].append('')

        result[1].append('')
        result[1].append('')

        result[1].append('')
        result[1].append('')

        for w in array:
            if l > 5 + self.radW * self.stolb:
                break
            answer = w[1]
            answer_is_true = w[2]
            if w[1] != 'klass':
                result[0].append(answer)
                result[1].append(answer_is_true)
            l += 1

        result[1].append(itog)

        return result